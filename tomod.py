# 2021-11-05
# SA
# licenc: MIT


###############################################
#
# Beállítások
#

in_file_name = 'minta.xlsx'
out_file_name = 'minta.csv'

# Az első sor ahol tanulók vannak a bemeneti fájlban
first_row = 1

# Itt beállítható milyen legyen egy sor:
#out_line_format = 'id, "first_name", "second_name", "user_name", om, "phone", "email"'
out_line_format = 'id, first_name, second_name, user_name, omid, phone, email'

#
###############################################


import openpyxl
import unidecode
 

wb_obj = openpyxl.load_workbook(in_file_name)
sheet_obj = wb_obj.active
max_col = sheet_obj.max_column
max_row = sheet_obj.max_row

def getFirstName(name):    
    (secondName,firstName) = name.split()
    return firstName

def getSecondName(name):    
    (secondName,firstName) = name.split()
    return secondName

def getUserName(name):    
    (secondName,firstName) = name.split()
    tmpName = secondName + "." + firstName
    userName = tmpName.lower()
    return unaccent(userName)

def unaccent(name):    
    unaccented_name = unidecode.unidecode(name)
    return unaccented_name

f = open(out_file_name, 'w')

for i in range(first_row, max_row + 1):
    id = sheet_obj.cell(row = i, column = 1).value
    full_name = sheet_obj.cell(row = i, column = 2).value
    om = sheet_obj.cell(row = i, column = 3).value
    phone = sheet_obj.cell(row = i, column = 4).value
    email = sheet_obj.cell(row = i, column = 5).value

    first_name = getFirstName(full_name)
    second_name = getSecondName(full_name)
    user_name = getUserName(full_name)
    
    line = out_line_format \
    .replace('id', str(id)) \
    .replace('first_name', first_name) \
    .replace('second_name', second_name) \
    .replace('user_name', user_name) \
    .replace('om', str(om)) \
    .replace('phone', str(phone)) \
    .replace('email', email)

    print(line)

    line = (
            str(id) + "," + 
            first_name + "," +
            second_name + "," +
            user_name + ", " +
            str(om) + ", " +
            str(phone) + ", " +
            email + "\n"
        )
    f.write(line)
f.close()

